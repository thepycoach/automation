import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
import string

# Leer achivo excel
archivo_excel = pd.read_excel('supermarket_sales.xlsx')
# Hacer tabla pivote y exportar a excel (sales_2021.xlsl)
tabla_pivote = archivo_excel.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
tabla_pivote.to_excel('sales_2021.xlsx', startrow=4, sheet_name='Report')

# cargar archivo excel creado para manipularlo con Python
wb = load_workbook('sales_2021.xlsx')
pestaña = wb['Report']

# referencias de filas/columnas
min_col = wb.active.min_column
max_col = wb.active.max_column
min_fila = wb.active.min_row
max_fila = wb.active.max_row

# creando gráficos/charts
barchart = BarChart()
data = Reference(pestaña, min_col=min_col+1, max_col=max_col, min_row=min_fila, max_row=max_fila)
categorias = Reference(pestaña, min_col=min_col, max_col=min_col, min_row=min_fila+1, max_row=max_fila)
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categorias)
pestaña.add_chart(barchart, 'B12')
barchart.title = 'Ventas'
barchart.style = 2

# creando abecedario para usar como referencia de columnas en Excel
abecedario = list(string.ascii_uppercase)
abecedario_excel = abecedario[0:max_col]

# aplicando formulas a celdas de Excel
for i in abecedario_excel:
    if i!='A':
        pestaña[f'{i}{max_fila+1}'] = f'=SUM({i}{min_fila+1}:{i}{max_fila})'
        pestaña[f'{i}{max_fila+1}'].style = 'Currency'
pestaña[f'{abecedario_excel[0]}{max_fila+1}'] = 'Total'

# Dando formato al reporte de Excel
pestaña['A1'] = 'Reporte'
pestaña['A2'] = '2021'
pestaña['A1'].font = Font('Arial', bold=True, size=20)
pestaña['A2'].font = Font('Arial', bold=True, size=12)

wb.save('sales_2021.xlsx')
