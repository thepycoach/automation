import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
import string
import sys, os

# ubicacion del archivo ejecutable
application_path = os.path.dirname(sys.executable)

def automatizar_excel(nombre_archivo):
    """Input sales_mes.xlsx / Output report_mes.xlsx"""
    # creando y exportando tabla pivote a Excel
    archivo_excel = pd.read_excel(f'{application_path}/{nombre_archivo}')
    tabla_pivote = archivo_excel.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
    mes_extension = nombre_archivo.split('_')[1]
    tabla_pivote.to_excel(f'{application_path}/report_{mes_extension}', startrow=4, sheet_name='Report')

    # leer archivo excel para manipularlo con python
    wb = load_workbook(f'{application_path}/report_{mes_extension}')
    pestaña = wb['Report']

    min_col = wb.active.min_column
    max_col = wb.active.max_column
    min_fila = wb.active.min_row
    max_fila = wb.active.max_row

    # chart
    barchart = BarChart()
    data = Reference(pestaña, min_col=min_col+1, max_col=max_col, min_row=min_fila, max_row=max_fila)
    categorias = Reference(pestaña, min_col=min_col, max_col=min_col, min_row=min_fila+1, max_row=max_fila)
    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categorias)
    pestaña.add_chart(barchart, 'B12')
    barchart.title = 'Ventas'
    barchart.style = 2

    # formulas Excel
    abecedario = list(string.ascii_uppercase)
    abecedario_excel = abecedario[0:max_col]
    for i in abecedario_excel:
        if i!='A':
            pestaña[f'{i}{max_fila+1}'] = f'=SUM({i}{min_fila+1}:{i}{max_fila})'
            pestaña[f'{i}{max_fila+1}'].style = 'Currency'

    pestaña[f'{abecedario_excel[0]}{max_fila+1}'] = 'Total'

    # formato
    pestaña['A1'] = 'Reporte'
    mes = mes_extension.split('.')[0]
    pestaña['A2'] = mes
    pestaña['A1'].font = Font('Arial', bold=True, size=20)
    pestaña['A2'].font = Font('Arial', bold=True, size=12)

    wb.save(f'{application_path}/report_{mes_extension}')
    return

# automatizar reporte 2021
nombre_archivo = input('Insert file name: ')
automatizar_excel(f'{nombre_archivo}.xlsx')
